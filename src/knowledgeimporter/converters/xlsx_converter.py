"""XLSX to chunking-safe Markdown converter."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from .base import BaseConverter, RawDocument, Section


class XlsxConverter(BaseConverter):
    """Converts XLSX files into chunking-safe Markdown (one section per data row)."""

    def extract(self, path: str) -> RawDocument:
        wb = openpyxl.load_workbook(path, data_only=True)
        sections: list[Section] = []
        all_text_parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            # Filter empty rows
            rows = [r for r in rows if any(c is not None for c in r)]
            if not rows:
                continue

            headers = [str(h) if h is not None else f"Spalte{i}" for i, h in enumerate(rows[0], start=1)]
            all_text_parts.extend(str(h) for h in headers)

            for idx, row in enumerate(rows[1:], start=2):
                kv: list[tuple[str, str]] = []
                for h, val in zip(headers, row, strict=False):
                    if val is not None:
                        s_val = str(val)
                        all_text_parts.append(s_val)
                        kv.append((h, s_val))
                if not kv:
                    continue
                first_val = kv[0][1] if kv else str(idx)
                title = f"{sheet_name} – Zeile {idx}: {first_val}"
                sections.append(Section(level=2, title=title, kv_pairs=kv))

        return RawDocument(
            source_path=path,
            source_type="xlsx",
            title=Path(path).stem,
            language="de",
            date=None,
            sections=sections,
            metadata={"sheets": wb.sheetnames},
            raw_text=" ".join(all_text_parts),
        )
