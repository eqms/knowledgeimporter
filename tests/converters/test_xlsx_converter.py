"""Tests for XlsxConverter — XLSX to chunking-safe Markdown."""

import tempfile

import openpyxl
import pytest


def create_xlsx(rows: list[list]) -> str:
    """Create minimal XLSX file for tests."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(f.name)
    return f.name


def test_xlsx_extracts_rows_as_sections():
    from knowledgeimporter.converters.xlsx_converter import XlsxConverter

    path = create_xlsx([
        ["Name", "Preis", "Bestand"],
        ["Artikel A", 4.50, 100],
        ["Artikel B", 9.90, 50],
    ])
    doc = XlsxConverter().extract(path)
    assert len(doc.sections) == 2


def test_xlsx_kv_uses_header_names():
    from knowledgeimporter.converters.xlsx_converter import XlsxConverter

    path = create_xlsx([
        ["Bezeichnung", "Preis"],
        ["Klebeband", 3.50],
    ])
    doc = XlsxConverter().extract(path)
    keys = [k for k, _ in doc.sections[0].kv_pairs]
    assert "Bezeichnung" in keys
    assert "Preis" in keys


def test_xlsx_multiple_sheets():
    from knowledgeimporter.converters.xlsx_converter import XlsxConverter

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Blatt1"
    ws1.append(["A", "B"])
    ws1.append([1, 2])
    ws2 = wb.create_sheet("Blatt2")
    ws2.append(["X", "Y"])
    ws2.append([10, 20])
    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(f.name)

    doc = XlsxConverter().extract(f.name)
    titles = [s.title for s in doc.sections]
    assert any("Blatt1" in t or "Blatt2" in t for t in titles)


def test_xlsx_markdown_has_frontmatter():
    from knowledgeimporter.converters.xlsx_converter import XlsxConverter

    path = create_xlsx([["X"], [1]])
    result = XlsxConverter().run(path)
    assert result.markdown_content.startswith("---")
    assert "quelle: xlsx" in result.markdown_content
